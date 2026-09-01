#!/usr/bin/python3
# Copyright (c) KMG. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
"""Collect system + container info and append a ``system`` sheet to the
sbk-charts xlsx.

The sheet uses a **column-per-attribute, row-per-host** layout. There is
exactly one row for every distinct system that participated in the run:

- One row for the **local** host if any ``sbk-yal`` instance ran (one row
  total, regardless of how many local sbk-yal instances).
- One row per **distinct remote node** if any ``sbk-gem-yal`` instance ran.
  Remote rows are populated by SSHing into each node (using ``gemuser`` /
  ``gempass`` / ``gemport`` from the instance YAML) and running a small
  POSIX shell script that prints key=value pairs.

Each row records OS, CPU, RAM, **Docker / Kubernetes container details**,
and which sbk-analytics instances ran on that system.
"""
from __future__ import annotations

import logging
import os
import platform
import shutil
import socket
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Iterable

import psutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .policy import RUNTIME_POLICY

log = logging.getLogger(__name__)
SYSTEM_INFO_POLICY = RUNTIME_POLICY.system_info
SSH_POLICY = RUNTIME_POLICY.ssh
DISPLAY_POLICY = RUNTIME_POLICY.display


# Columns of the system sheet, in order.
SYSTEM_COLUMNS: list[str] = list(SYSTEM_INFO_POLICY.columns)


# ---------------------------------------------------------------------------
# Local system info
# ---------------------------------------------------------------------------


def _cpu_brand() -> str:
    """Best-effort CPU brand/model string across platforms."""
    sysname = platform.system()
    try:
        if sysname == SYSTEM_INFO_POLICY.linux_platform:
            try:
                with open(SYSTEM_INFO_POLICY.cpu_info_file) as f:
                    for line in f:
                        if line.lower().startswith("model name"):
                            return line.split(":", 1)[1].strip()
            except OSError:
                pass
            if shutil.which(SYSTEM_INFO_POLICY.lscpu_command[0]):
                out = subprocess.run(
                    list(SYSTEM_INFO_POLICY.lscpu_command),
                    capture_output=True, text=True,
                    timeout=SYSTEM_INFO_POLICY.local_command_timeout_s,
                ).stdout
                for line in out.splitlines():
                    if line.lower().startswith("model name"):
                        return line.split(":", 1)[1].strip()
        elif sysname == SYSTEM_INFO_POLICY.macos_platform:
            out = subprocess.run(
                list(SYSTEM_INFO_POLICY.macos_cpu_command),
                capture_output=True, text=True,
                timeout=SYSTEM_INFO_POLICY.local_command_timeout_s,
            ).stdout.strip()
            if out:
                return out
    except Exception as e:
        log.debug("cpu brand lookup failed: %s", e)
    return platform.processor() or DISPLAY_POLICY.unknown_value


def _container_info() -> dict[str, str]:
    """Detect Docker / Kubernetes container details for the current process."""
    runtime = DISPLAY_POLICY.absent_value
    container_id = ""
    k8s_pod = ""
    k8s_namespace = ""

    if Path(SYSTEM_INFO_POLICY.docker_environment_file).exists():
        runtime = SYSTEM_INFO_POLICY.docker_runtime
    if os.environ.get(SYSTEM_INFO_POLICY.kubernetes_service_environment):
        runtime = SYSTEM_INFO_POLICY.kubernetes_runtime
    try:
        cgroup = Path(SYSTEM_INFO_POLICY.process_cgroup_file).read_text()
        if "/kubepods" in cgroup or "/kubelet" in cgroup:
            runtime = SYSTEM_INFO_POLICY.kubernetes_runtime
        elif (
            runtime == DISPLAY_POLICY.absent_value
            and ("/docker" in cgroup or "/containerd" in cgroup)
        ):
            runtime = SYSTEM_INFO_POLICY.docker_runtime
    except OSError:
        pass

    if runtime != DISPLAY_POLICY.absent_value:
        try:
            for line in Path(
                SYSTEM_INFO_POLICY.self_cgroup_file
            ).read_text().splitlines():
                last = line.rstrip().split("/")[-1]
                if last:
                    container_id = last[
                        :SYSTEM_INFO_POLICY.container_id_characters
                    ]
                    break
        except OSError:
            pass

    if runtime == "kubernetes":
        k8s_pod = (
            os.environ.get("POD_NAME")
            or os.environ.get("HOSTNAME")
            or socket.gethostname()
        )
        try:
            ns_path = Path("/var/run/secrets/kubernetes.io/serviceaccount/namespace")
            if ns_path.is_file():
                k8s_namespace = ns_path.read_text().strip()
        except OSError:
            pass

    return {
        SYSTEM_INFO_POLICY.container_runtime_column: runtime,
        SYSTEM_INFO_POLICY.container_id_column: container_id,
        SYSTEM_INFO_POLICY.kubernetes_pod_column: k8s_pod,
        SYSTEM_INFO_POLICY.kubernetes_namespace_column: k8s_namespace,
    }


def collect_local_system_info() -> dict[str, str]:
    """Return system + container info for the local host."""
    vm = psutil.virtual_memory()
    freq = None
    try:
        f = psutil.cpu_freq()
        if f and f.max:
            freq = f"{f.max:.0f}"
    except Exception:
        pass

    info: dict[str, str] = {
        SYSTEM_INFO_POLICY.hostname_column: socket.gethostname(),
        SYSTEM_INFO_POLICY.operating_system_column: (
            f"{platform.system()} {platform.release()}"
        ),
        SYSTEM_INFO_POLICY.operating_system_version_column: platform.version(),
        SYSTEM_INFO_POLICY.architecture_column: platform.machine(),
        SYSTEM_INFO_POLICY.cpu_model_column: _cpu_brand(),
        SYSTEM_INFO_POLICY.physical_cpus_column: str(
            psutil.cpu_count(logical=False) or ""
        ),
        SYSTEM_INFO_POLICY.logical_cpus_column: str(
            psutil.cpu_count(logical=True) or ""
        ),
        SYSTEM_INFO_POLICY.cpu_mhz_column: freq or "",
        SYSTEM_INFO_POLICY.total_ram_column: (
            f"{vm.total / (DISPLAY_POLICY.bytes_per_kibibyte ** 3):.2f}"
        ),
        SYSTEM_INFO_POLICY.available_ram_column: (
            f"{vm.available / (DISPLAY_POLICY.bytes_per_kibibyte ** 3):.2f}"
        ),
    }
    info.update(_container_info())
    info[SYSTEM_INFO_POLICY.collected_at_column] = datetime.now().isoformat(
        timespec="seconds"
    )
    info[SYSTEM_INFO_POLICY.status_column] = RUNTIME_POLICY.cli.success_status
    return info


# ---------------------------------------------------------------------------
# Remote system info (over SSH)
# ---------------------------------------------------------------------------


# A small POSIX-shell probe that prints key=value pairs to stdout. Designed
# to run on bare-metal Linux, Docker containers, and Kubernetes pods alike.
_REMOTE_INFO_SCRIPT = r"""
set +e
echo "hostname=$(hostname 2>/dev/null)"
echo "os=$(uname -s 2>/dev/null) $(uname -r 2>/dev/null)"
echo "os_version=$(uname -v 2>/dev/null)"
echo "arch=$(uname -m 2>/dev/null)"
echo "logical_cpus=$(nproc 2>/dev/null)"
echo "physical_cpus=$(lscpu 2>/dev/null | awk -F: '/^Socket\(s\)/{gsub(/ /,"",$2); print $2}')"
echo "cpu_model=$(awk -F: '/^model name/{sub(/^ +/,"",$2); print $2; exit}' /proc/cpuinfo 2>/dev/null)"
echo "cpu_mhz=$(awk -F: '/^cpu MHz/{gsub(/ /,"",$2); print $2; exit}' /proc/cpuinfo 2>/dev/null)"
mem_total_kb=$(awk '/^MemTotal:/{print $2}' /proc/meminfo 2>/dev/null)
mem_avail_kb=$(awk '/^MemAvailable:/{print $2}' /proc/meminfo 2>/dev/null)
echo "total_ram_kb=${mem_total_kb:-}"
echo "avail_ram_kb=${mem_avail_kb:-}"
runtime=none
if [ -f /.dockerenv ]; then runtime=docker; fi
if [ -n "$KUBERNETES_SERVICE_HOST" ]; then runtime=kubernetes; fi
if grep -q "/kubepods\|/kubelet" /proc/1/cgroup 2>/dev/null; then runtime=kubernetes; fi
if [ "$runtime" = "none" ] && grep -q "/docker\|/containerd" /proc/1/cgroup 2>/dev/null; then runtime=docker; fi
echo "container_runtime=$runtime"
container_id=$(awk -F/ '{ for (i=NF; i>=1; i--) if ($i != "") { print $i; exit } }' /proc/self/cgroup 2>/dev/null | head -c __CONTAINER_ID_CHARACTERS__)
echo "container_id=${container_id}"
echo "k8s_pod=${POD_NAME:-${HOSTNAME:-}}"
ns=""
if [ -r /var/run/secrets/kubernetes.io/serviceaccount/namespace ]; then
  ns=$(cat /var/run/secrets/kubernetes.io/serviceaccount/namespace 2>/dev/null)
fi
echo "k8s_namespace=${ns}"
""".replace(
    "__CONTAINER_ID_CHARACTERS__",
    str(SYSTEM_INFO_POLICY.container_id_characters),
)


def collect_remote_system_info(
    node: str,
    *,
    user: str = "",
    password: str = "",
    port: int = SSH_POLICY.default_port,
    timeout_s: float = SSH_POLICY.system_info_command_timeout_s,
) -> dict[str, str]:
    """Return system + container info for a remote node, via SSH.

    If SSH or the probe fails, returns a dict with ``Status`` set to a
    short error description; the row will still be added to the sheet with
    whatever info we already know (just the node name in that case).
    """
    target = f"{user}@{node}" if user else node
    ssh_args = [
        SSH_POLICY.ssh_command,
        SSH_POLICY.port_option, str(int(port)),
        *SSH_POLICY.host_key_arguments,
        SSH_POLICY.option_flag,
        SSH_POLICY.batch_mode_option + "=" + (
            SSH_POLICY.disabled_value
            if password else SSH_POLICY.enabled_value
        ),
        SSH_POLICY.option_flag,
        f"{SSH_POLICY.connect_timeout_option}={SSH_POLICY.connect_timeout_s}",
        target,
        SSH_POLICY.remote_shell_command,
    ]
    have_sshpass = (
        bool(password)
        and shutil.which(SSH_POLICY.sshpass_command) is not None
    )
    if password and not have_sshpass:
        log.warning(
            "remote system info: 'sshpass' not on PATH but gempass is set for "
            "%s; attempting key-based ssh which may fail", node,
        )
    if have_sshpass:
        env = {**os.environ, SSH_POLICY.sshpass_environment: password}
        cmd = [
            SSH_POLICY.sshpass_command,
            SSH_POLICY.sshpass_environment_option,
            *ssh_args,
        ]
    else:
        env = os.environ.copy()
        cmd = ssh_args

    try:
        proc = subprocess.run(
            cmd, env=env,
            input=_REMOTE_INFO_SCRIPT,
            capture_output=True, text=True, timeout=timeout_s,
        )
    except subprocess.TimeoutExpired:
        return {
            SYSTEM_INFO_POLICY.status_column:
                f"ssh timeout after {int(timeout_s)}s"
        }
    except OSError as e:
        return {SYSTEM_INFO_POLICY.status_column: f"ssh error: {e}"}

    if proc.returncode != 0:
        err = (proc.stderr or "").strip().splitlines()[-1:] or [""]
        return {
            SYSTEM_INFO_POLICY.status_column: (
                f"ssh rc={proc.returncode}: "
                f"{err[0][:DISPLAY_POLICY.system_info_tail_characters]}"
            )
        }

    raw: dict[str, str] = {}
    for line in proc.stdout.splitlines():
        line = line.rstrip()
        if "=" not in line:
            continue
        k, v = line.split("=", 1)
        raw[k.strip()] = v.strip()

    def _bytes_to_gib(kb_str: str) -> str:
        try:
            return (
                f"{int(kb_str) / (DISPLAY_POLICY.bytes_per_kibibyte ** 2):.2f}"
            )
        except (TypeError, ValueError):
            return ""

    info: dict[str, str] = {
        SYSTEM_INFO_POLICY.hostname_column: raw.get(
            SYSTEM_INFO_POLICY.remote_hostname_field, ""
        ),
        SYSTEM_INFO_POLICY.operating_system_column: raw.get(
            SYSTEM_INFO_POLICY.remote_os_field, ""
        ),
        SYSTEM_INFO_POLICY.operating_system_version_column: raw.get(
            SYSTEM_INFO_POLICY.remote_os_version_field, ""
        ),
        SYSTEM_INFO_POLICY.architecture_column: raw.get(
            SYSTEM_INFO_POLICY.remote_architecture_field, ""
        ),
        SYSTEM_INFO_POLICY.cpu_model_column: raw.get(
            SYSTEM_INFO_POLICY.remote_cpu_model_field, ""
        ),
        SYSTEM_INFO_POLICY.physical_cpus_column: raw.get(
            SYSTEM_INFO_POLICY.remote_physical_cpus_field, ""
        ),
        SYSTEM_INFO_POLICY.logical_cpus_column: raw.get(
            SYSTEM_INFO_POLICY.remote_logical_cpus_field, ""
        ),
        SYSTEM_INFO_POLICY.cpu_mhz_column: raw.get(
            SYSTEM_INFO_POLICY.remote_cpu_mhz_field, ""
        ),
        SYSTEM_INFO_POLICY.total_ram_column: _bytes_to_gib(raw.get(
            SYSTEM_INFO_POLICY.remote_total_ram_kb_field, ""
        )),
        SYSTEM_INFO_POLICY.available_ram_column: _bytes_to_gib(raw.get(
            SYSTEM_INFO_POLICY.remote_available_ram_kb_field, ""
        )),
        SYSTEM_INFO_POLICY.container_runtime_column: raw.get(
            SYSTEM_INFO_POLICY.remote_container_runtime_field, "none"
        ),
        SYSTEM_INFO_POLICY.container_id_column: raw.get(
            SYSTEM_INFO_POLICY.remote_container_id_field, ""
        ),
        SYSTEM_INFO_POLICY.kubernetes_pod_column: raw.get(
            SYSTEM_INFO_POLICY.remote_kubernetes_pod_field, ""
        ),
        SYSTEM_INFO_POLICY.kubernetes_namespace_column: raw.get(
            SYSTEM_INFO_POLICY.remote_kubernetes_namespace_field, ""
        ),
        SYSTEM_INFO_POLICY.collected_at_column: datetime.now().isoformat(
            timespec="seconds"
        ),
        SYSTEM_INFO_POLICY.status_column: RUNTIME_POLICY.cli.success_status,
    }
    return info


# ---------------------------------------------------------------------------
# Sheet writer
# ---------------------------------------------------------------------------


def append_system_sheet(
    xlsx_path: Path,
    *,
    sources: Iterable[dict],
    sheet_name: str = SYSTEM_INFO_POLICY.default_sheet_name,
) -> None:
    """Append (or replace) the ``system`` sheet on the xlsx with one row per
    distinct system.

    Each entry in ``sources`` describes one system::

        {"kind": "local", "instances": ["file", "rocksdb"]}
        {"kind": "remote", "node": "n1", "user": "kmg", "password": "pw",
         "port": 22, "instances": ["kafka"]}

    A single fallback row (``Source="local"``, no instances) is written if
    ``sources`` is empty -- that way the sheet always has at least one data
    row describing the host that produced the report.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"xlsx not found: {xlsx_path}")

    sources = list(sources)
    if not sources:
        sources = [{
            SYSTEM_INFO_POLICY.source_kind_field:
                SYSTEM_INFO_POLICY.local_source,
            SYSTEM_INFO_POLICY.source_instances_field: [],
        }]

    wb = load_workbook(xlsx_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(SYSTEM_COLUMNS)

    for src in sources:
        kind = src.get(
            SYSTEM_INFO_POLICY.source_kind_field,
            SYSTEM_INFO_POLICY.local_source,
        )
        instances = ", ".join(
            src.get(SYSTEM_INFO_POLICY.source_instances_field) or []
        )
        if kind == SYSTEM_INFO_POLICY.remote_source:
            node = src[SYSTEM_INFO_POLICY.source_node_field]
            source_label = f"{SYSTEM_INFO_POLICY.remote_source}: {node}"
            info = collect_remote_system_info(
                node,
                user=src.get(SYSTEM_INFO_POLICY.source_user_field, ""),
                password=src.get(
                    SYSTEM_INFO_POLICY.source_password_field, ""
                ),
                port=int(src.get(
                    SYSTEM_INFO_POLICY.source_port_field,
                    SSH_POLICY.default_port,
                )),
            )
        else:
            source_label = SYSTEM_INFO_POLICY.local_source
            info = collect_local_system_info()

        row = [source_label, instances] + [
            info.get(col, "") for col in SYSTEM_COLUMNS[2:]
        ]
        ws.append(row)

    # Column widths
    for idx, width in enumerate(SYSTEM_INFO_POLICY.column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Freeze the header row.
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    log.info(
        "appended '%s' sheet to %s (%d row%s)",
        sheet_name, xlsx_path, len(sources), "" if len(sources) == 1 else "s",
    )
